from datetime import datetime, timedelta

from django.http import HttpResponse
from django.shortcuts import render

from apps.data.models import Conferencia, Dependente, Inscricao
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def gen_report_cracha(request, pk):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    conferencia = Conferencia.objects.get(pk=pk)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-crachas.xlsx'.format(
        date=conferencia.titulo,
    )
    workbook = Workbook()
    
    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Inscrições'

    # Define the titles for columns
    columns = [
        'nickname',
        'cidade',
        'nickname',
        'idade',
    ]
    row_num = 1

    inscricao_qs = Inscricao.objects.filter(conferencia=conferencia).order_by('cidade')

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for inscr in inscricao_qs:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            inscr.nome_cracha,
            inscr.cidade,
            "",
            inscr.idade,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

        for dep in Dependente.objects.filter(inscricao=inscr).order_by('-idade'):
            row_num += 1
        
            row = [
                dep.nome_cracha,
                dep.inscricao.cidade,
                dep.inscricao.nome_cracha,
                dep.idade,
            ]
            
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

    column_letter = get_column_letter(1)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 35

    column_letter = get_column_letter(2)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 35

    column_letter = get_column_letter(3)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 35

    column_letter = get_column_letter(4)
    column_dimensions = worksheet.column_dimensions[column_letter]
    column_dimensions.width = 10


    workbook.save(response)

    return response
